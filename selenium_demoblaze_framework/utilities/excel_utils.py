# utilities/excel_utils.py

import openpyxl
import pandas as pd
import os


class ExcelUtils:
    @staticmethod
    def read_excel_data(file_path, sheet_name=None):
        """
        Read data from an Excel file using openpyxl.
        Returns a list of dictionaries, where each dictionary represents a row with headers as keys.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        data = []
        headers = [cell.value for cell in sheet[1]]  # First row as headers

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # Skip completely empty rows
                data.append(dict(zip(headers, row)))

        workbook.close()
        return data

    @staticmethod
    def write_excel_data(file_path, data, sheet_name='Sheet1'):
        """
        Write a list of dictionaries to an Excel file using openpyxl.
        Each dictionary represents a row; keys are used as column headers.
        """
        if not data:
            raise ValueError("No data provided to write to Excel.")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data rows
        for row_num, row_data in enumerate(data, start=2):
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row_num, column=col_num, value=row_data.get(header, ''))

        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)

        workbook.save(file_path)
        workbook.close()

    @staticmethod
    def read_with_pandas(file_path, sheet_name=None):
        """
        Read Excel file using pandas (supports .xlsx, .xls).
        Returns list of dictionaries.
        Useful for handling formulas, mixed data types, or large files.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        return df.to_dict('records')

    @staticmethod
    def update_cell(file_path, sheet_name, row, column, value):
        """
        Update a specific cell in an existing Excel file.
        Parameters:
            file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet (optional; uses active sheet if None).
            row (int): Row number (1-based index).
            column (int or str): Column number (1-based) or letter (e.g., 'A').
            value: The value to write into the cell.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        sheet.cell(row=row, column=column, value=value)
        workbook.save(file_path)
        workbook.close()